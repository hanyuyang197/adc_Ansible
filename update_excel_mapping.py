#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import glob

def find_py_module_mappings(library_dir):
    """从Python模块文件中查找action到模块名的映射"""
    py_mappings = {}
    duplicate_actions = {}
    
    # 遍历library目录下的所有.py文件
    py_files = glob.glob(os.path.join(library_dir, "*.py"))
    
    for py_file in py_files:
        with open(py_file, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            
        # 查找所有包含action=的URL模式
        action_pattern = r'&action=([\w\.]+)"'
        matches = re.findall(action_pattern, content)
        
        for action in matches:
            # 提取模块名（去掉adc_前缀和.py后缀）
            module_name = os.path.basename(py_file).replace('.py', '')
            
            if action in py_mappings:
                # 发现重复的action
                if action not in duplicate_actions:
                    duplicate_actions[action] = []
                duplicate_actions[action].append(py_file)
                duplicate_actions[action].append(py_mappings[action]['file'])
            else:
                # 提取action名（从URL的最后部分）
                action_parts = action.split('.')
                action_name = '_'.join(action_parts)
                
                py_mappings[action] = {
                    'module': module_name,
                    'action': action_name,
                    'file': py_file
                }
    
    return py_mappings, duplicate_actions

def find_yml_mappings(playbooks_dir):
    """从YAML文件中查找action到模块名的映射"""
    yml_mappings = {}
    
    # 遍历playbooksNew目录下的所有.yml文件
    yml_files = glob.glob(os.path.join(playbooks_dir, "*.yml"))
    
    for yml_file in yml_files:
        with open(yml_file, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            
        # 查找API action注释
        action_pattern = r'# API action: ([\w\.]+)'
        matches = re.findall(action_pattern, content)
        
        for action in matches:
            # 查找模块名（在adc_模块名:行中）
            module_pattern = r'adc_(\w+):'
            module_match = re.search(module_pattern, content)
            
            if module_match:
                module_name = f"adc_{module_match.group(1)}"
                
                # 提取action名（从文件名中）
                filename = os.path.basename(yml_file).replace('.yml', '')
                action_name = filename.replace(f"adc_{module_match.group(1)}_", "")
                
                yml_mappings[action] = {
                    'module': module_name,
                    'action': action_name,
                    'file': yml_file
                }
    
    return yml_mappings

def update_excel_file(excel_file, py_mappings, yml_mappings, duplicate_actions):
    """更新Excel文件"""
    
    # 加载Excel文件
    workbook = load_workbook(excel_file)
    
    # 获取所有sheet名
    sheet_names = workbook.sheetnames
    
    # 记录更新的内容
    updates = []
    
    for sheet_name in sheet_names:
        if sheet_name not in ['slb', 'system', 'network']:  # 根据实际sheet名调整
            continue
            
        sheet = workbook[sheet_name]
        
        # 读取B列（URL）和C列（Ansible对应）
        for row in range(2, sheet.max_row + 1):  # 从第2行开始（跳过标题行）
            url_cell = sheet.cell(row=row, column=2)  # B列
            ansible_cell = sheet.cell(row=row, column=3)  # C列
            module_cell = sheet.cell(row=row, column=6)  # F列
            action_cell = sheet.cell(row=row, column=7)  # G列
            
            url_value = url_cell.value
            
            if not url_value:
                continue
            
            # 检查是否已经人工校对过（C列包含playbooksNew）
            ansible_value = ansible_cell.value if ansible_cell.value else ""
            
            if 'playbooksNew' in str(ansible_value):
                # 已经人工校对过，只需要更新F列和G列
                if url_value in py_mappings:
                    mapping = py_mappings[url_value]
                    module_cell.value = mapping['module']
                    action_cell.value = mapping['action']
                    updates.append(f"Sheet {sheet_name}, 行 {row}: 更新F/G列 - {url_value}")
            else:
                # 需要更新C、F、G列
                if url_value in py_mappings and url_value in yml_mappings:
                    py_mapping = py_mappings[url_value]
                    yml_mapping = yml_mappings[url_value]
                    
                    # 更新C列（Ansible对应）
                    ansible_cell.value = f"playbooksNew/{yml_mapping['file']}"
                    
                    # 更新F列（模块名）
                    module_cell.value = py_mapping['module']
                    
                    # 更新G列（action名）
                    action_cell.value = py_mapping['action']
                    
                    updates.append(f"Sheet {sheet_name}, 行 {row}: 更新C/F/G列 - {url_value}")
                elif url_value in py_mappings:
                    # 只有py映射，没有yml映射
                    mapping = py_mappings[url_value]
                    module_cell.value = mapping['module']
                    action_cell.value = mapping['action']
                    updates.append(f"Sheet {sheet_name}, 行 {row}: 更新F/G列（只有py映射）- {url_value}")
    
    # 保存Excel文件
    workbook.save(excel_file)
    
    return updates

def create_duplicate_report(duplicate_actions):
    """创建重复action的报告"""
    report_content = "# 重复的URL Action报告\n\n"
    
    if not duplicate_actions:
        report_content += "没有发现重复的URL action。\n"
    else:
        report_content += "以下URL action在多个Python模块文件中出现，需要人工筛选：\n\n"
        
        for action, files in duplicate_actions.items():
            report_content += f"## Action: {action}\n"
            report_content += "出现在以下文件中：\n"
            for file_path in files:
                report_content += f"- {file_path}\n"
            report_content += "\n"
    
    return report_content

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 文件路径
    library_dir = os.path.join(base_dir, "library")
    playbooks_dir = os.path.join(base_dir, "playbooksNew")
    excel_file = os.path.join(base_dir, "ansible_2026年1月6日.xlsx")
    
    print("开始分析Python模块文件...")
    py_mappings, duplicate_actions = find_py_module_mappings(library_dir)
    print(f"找到 {len(py_mappings)} 个Python模块映射")
    
    print("开始分析YAML文件...")
    yml_mappings = find_yml_mappings(playbooks_dir)
    print(f"找到 {len(yml_mappings)} 个YAML文件映射")
    
    print("开始更新Excel文件...")
    updates = update_excel_file(excel_file, py_mappings, yml_mappings, duplicate_actions)
    print(f"完成 {len(updates)} 个更新")
    
    # 创建重复action报告
    print("创建重复action报告...")
    report_content = create_duplicate_report(duplicate_actions)
    report_file = os.path.join(base_dir, "duplicate_actions_report.md")
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report_content)
    
    print(f"重复action报告已保存到: {report_file}")
    
    # 输出更新摘要
    print("\n更新摘要:")
    for update in updates:
        print(f"  {update}")
    
    if duplicate_actions:
        print(f"\n发现 {len(duplicate_actions)} 个重复的URL action，请查看报告文件进行人工筛选")

if __name__ == "__main__":
    main()